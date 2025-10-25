"""
RCM (Risk and Control Matrix) Routes
ELC/TLC/ITGC RCM 등록 및 관리
"""

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, session
import sys
import os

# Sentinel 루트 경로를 Python path에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from common.auth import (
    login_required, admin_required, get_current_user, get_user_rcms,
    has_rcm_access, get_rcm_details, get_rcm_info, create_rcm,
    save_rcm_details, grant_rcm_access, log_user_activity, get_db
)

bp_rcm = Blueprint('rcm', __name__, url_prefix='/rcm')

def get_user_info():
    """현재 로그인한 사용자 정보 반환"""
    return get_current_user()

def is_logged_in():
    """로그인 상태 확인"""
    return 'user_info' in session

# RCM 목록 조회
@bp_rcm.route('/')
@bp_rcm.route('/list')
@login_required
def rcm_list():
    """RCM 목록 페이지 (ELC/TLC/ITGC 전체)"""
    user_info = get_user_info()

    # 사용자가 접근 가능한 모든 RCM 조회
    all_rcms = get_user_rcms(user_info['user_id'])

    # 카테고리별로 분류
    rcms_by_category = {
        'ELC': [rcm for rcm in all_rcms if rcm['control_category'] == 'ELC'],
        'TLC': [rcm for rcm in all_rcms if rcm['control_category'] == 'TLC'],
        'ITGC': [rcm for rcm in all_rcms if rcm['control_category'] == 'ITGC']
    }

    log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 목록 조회', '/rcm/list',
                     request.remote_addr, request.headers.get('User-Agent'),
                     {'total_rcms': len(all_rcms)})

    return render_template('rcm/rcm_list.html',
                         rcms_by_category=rcms_by_category,
                         all_rcms=all_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# 카테고리별 RCM 목록
@bp_rcm.route('/<category>')
@login_required
def rcm_category(category):
    """카테고리별 RCM 목록 페이지 (ELC, TLC, ITGC)"""
    if category not in ['ELC', 'TLC', 'ITGC']:
        flash('잘못된 카테고리입니다.')
        return redirect(url_for('rcm.rcm_list'))

    user_info = get_user_info()

    # 해당 카테고리의 RCM만 조회
    rcms = get_user_rcms(user_info['user_id'], control_category=category)

    category_names = {
        'ELC': 'Entity Level Controls (전사적 통제)',
        'TLC': 'Transaction Level Controls (거래 수준 통제)',
        'ITGC': 'IT General Controls (IT 일반 통제)'
    }

    log_user_activity(user_info, 'PAGE_ACCESS', f'{category} RCM 목록 조회',
                     f'/rcm/{category}', request.remote_addr,
                     request.headers.get('User-Agent'),
                     {'category': category, 'count': len(rcms)})

    return render_template('rcm/rcm_category.html',
                         category=category,
                         category_name=category_names[category],
                         rcms=rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# RCM 상세 조회
@bp_rcm.route('/<int:rcm_id>/view')
@login_required
def rcm_view(rcm_id):
    """RCM 상세 조회 페이지"""
    user_info = get_user_info()

    # 접근 권한 확인
    if not has_rcm_access(user_info['user_id'], rcm_id):
        flash('해당 RCM에 대한 접근 권한이 없습니다.')
        return redirect(url_for('rcm.rcm_list'))

    # RCM 기본 정보
    rcm_info = get_rcm_info(rcm_id)
    if not rcm_info:
        flash('RCM을 찾을 수 없습니다.')
        return redirect(url_for('rcm.rcm_list'))

    # RCM 상세 데이터 (통제 목록)
    rcm_details = get_rcm_details(rcm_id)

    log_user_activity(user_info, 'RCM_VIEW', f'RCM 상세 조회 - {rcm_info["rcm_name"]}',
                     f'/rcm/{rcm_id}/view', request.remote_addr,
                     request.headers.get('User-Agent'),
                     {'rcm_id': rcm_id, 'control_count': len(rcm_details)})

    return render_template('rcm/rcm_view.html',
                         rcm_info=rcm_info,
                         rcm_details=rcm_details,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# RCM 업로드 페이지
@bp_rcm.route('/upload')
@admin_required
def rcm_upload():
    """RCM 업로드 페이지 (관리자 전용)"""
    user_info = get_user_info()

    # 모든 사용자 목록 조회
    db = get_db()
    users = db.execute('''
        SELECT user_id, user_name, user_email, company_name
        FROM ca_user
        WHERE effective_end_date IS NULL OR effective_end_date > CURRENT_TIMESTAMP
        ORDER BY company_name, user_name
    ''').fetchall()
    users_list = [dict(user) for user in users]

    log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 업로드 페이지 접근',
                     '/rcm/upload', request.remote_addr,
                     request.headers.get('User-Agent'))

    return render_template('rcm/rcm_upload.html',
                         users=users_list,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)

# RCM 업로드 처리 (1단계: Excel 파일 업로드 및 컬럼 매핑)
@bp_rcm.route('/process_upload', methods=['POST'])
@admin_required
def rcm_process_upload():
    """Excel 파일 업로드 처리 (1단계)"""
    user_info = get_user_info()

    try:
        rcm_name = request.form.get('rcm_name', '').strip()
        upload_mode = request.form.get('upload_mode', 'individual').strip()
        control_category = request.form.get('control_category', '').strip()
        description = request.form.get('description', '').strip()
        target_user_id = request.form.get('target_user_id', '').strip()

        # 유효성 검사
        if not rcm_name:
            return jsonify({'success': False, 'message': 'RCM명은 필수입니다.'})

        # 개별 업로드 모드에서만 카테고리 체크
        if upload_mode == 'individual' and control_category not in ['ELC', 'TLC', 'ITGC']:
            return jsonify({'success': False, 'message': '통제 카테고리를 선택해주세요.'})

        if not target_user_id:
            return jsonify({'success': False, 'message': '대상 사용자를 선택해주세요.'})

        target_user_id = int(target_user_id)

        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'message': 'Excel 파일을 선택해주세요.'})

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Excel 파일을 선택해주세요.'})

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'})

        # RCM 생성 (개별 업로드 모드에서만)
        if upload_mode == 'individual':
            rcm_id = create_rcm(rcm_name, control_category, description, target_user_id, file.filename)
        else:
            rcm_id = None  # 통합 모드에서는 나중에 카테고리별로 생성

        # Excel 파일 읽기
        from openpyxl import load_workbook
        import tempfile
        import json

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        try:
            # Excel 파일 읽기
            workbook = load_workbook(temp_file.name)

            # RCM 시트 선택
            if 'RCM' in workbook.sheetnames:
                sheet = workbook['RCM']
            else:
                sheet = workbook.active

            # 헤더 추출
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else '')

            # 샘플 데이터 추출 (최대 5행)
            sample_data = []
            max_sample_rows = min(6, sheet.max_row)
            for row_num in range(2, max_sample_rows + 1):
                row_data = []
                for col_num in range(1, len(headers) + 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                sample_data.append(row_data)

            # 자동 매핑 수행
            auto_mapping = perform_auto_mapping(headers)

            if upload_mode == 'integrated':
                # 통합 업로드 모드: 카테고리 컬럼 찾기
                category_col_idx = find_category_column(headers)
                if category_col_idx is None:
                    return jsonify({
                        'success': False,
                        'message': '통합 업로드 모드에서는 "카테고리" 또는 "category" 컬럼이 필요합니다.'
                    })

                # 카테고리별로 RCM 생성 및 데이터 저장
                categories_data = {'ELC': [], 'TLC': [], 'ITGC': []}
                rcm_ids = {}

                # 데이터를 카테고리별로 분류
                for row_num in range(2, sheet.max_row + 1):
                    category_value = str(sheet.cell(row=row_num, column=category_col_idx + 1).value or '').strip().upper()

                    if category_value not in ['ELC', 'TLC', 'ITGC']:
                        continue  # 잘못된 카테고리는 건너뛰기

                    control = {}
                    for db_field, col_idx in auto_mapping.items():
                        cell_value = sheet.cell(row=row_num, column=col_idx + 1).value
                        control[db_field] = str(cell_value) if cell_value is not None else ''

                    if control.get('control_code'):
                        categories_data[category_value].append(control)

                # 각 카테고리별로 RCM 생성
                total_controls = 0
                for category, controls in categories_data.items():
                    if controls:  # 해당 카테고리에 데이터가 있는 경우만
                        category_rcm_name = f"{rcm_name} - {category}"
                        category_rcm_id = create_rcm(category_rcm_name, category, description,
                                                    target_user_id, file.filename)
                        rcm_ids[category] = category_rcm_id

                        # 데이터 저장
                        save_rcm_details(category_rcm_id, controls)
                        grant_rcm_access(target_user_id, category_rcm_id, user_info['user_id'], 'READ')
                        total_controls += len(controls)

                log_user_activity(user_info, 'RCM_UPLOAD_COMPLETE',
                                f'RCM 통합 업로드 완료 - {rcm_name}',
                                '/rcm/process_upload', request.remote_addr,
                                request.headers.get('User-Agent'),
                                {'rcm_name': rcm_name, 'mode': 'integrated',
                                 'categories': list(rcm_ids.keys()), 'total_controls': total_controls})

                return jsonify({
                    'success': True,
                    'message': f'RCM이 성공적으로 업로드되었습니다. (총 통제 수: {total_controls})',
                    'rcm_id': list(rcm_ids.values())[0] if rcm_ids else None,  # 첫 번째 RCM ID 반환
                    'rcm_ids': rcm_ids,
                    'controls_count': total_controls
                })
            else:
                # 개별 업로드 모드 (기존 로직)
                controls_data = []
                for row_num in range(2, sheet.max_row + 1):
                    control = {}
                    for db_field, col_idx in auto_mapping.items():
                        cell_value = sheet.cell(row=row_num, column=col_idx + 1).value
                        control[db_field] = str(cell_value) if cell_value is not None else ''

                    # 필수 필드 확인
                    if control.get('control_code'):
                        controls_data.append(control)

                # RCM 상세 데이터 저장
                if controls_data:
                    save_rcm_details(rcm_id, controls_data)

                # 사용자에게 RCM 접근 권한 부여
                grant_rcm_access(target_user_id, rcm_id, user_info['user_id'], 'READ')

                log_user_activity(user_info, 'RCM_UPLOAD_COMPLETE',
                                f'RCM 업로드 완료 - {rcm_name} ({control_category})',
                                '/rcm/process_upload', request.remote_addr,
                                request.headers.get('User-Agent'),
                                {'rcm_id': rcm_id, 'category': control_category, 'controls_count': len(controls_data)})

                return jsonify({
                    'success': True,
                    'message': f'RCM이 성공적으로 업로드되었습니다. (통제 수: {len(controls_data)})',
                    'rcm_id': rcm_id,
                    'controls_count': len(controls_data)
                })

        finally:
            # 임시 파일 삭제
            os.unlink(temp_file.name)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'업로드 중 오류가 발생했습니다: {str(e)}'}), 500

def find_category_column(headers):
    """카테고리 컬럼 찾기 (통합 업로드용)"""
    category_keywords = ['카테고리', 'category', '구분', 'type', 'class']

    for idx, header in enumerate(headers):
        header_lower = str(header).lower().strip()
        if any(keyword.lower() in header_lower for keyword in category_keywords):
            return idx
    return None

def perform_auto_mapping(headers):
    """Excel 헤더 자동 매핑"""
    mapping = {}

    # 매핑 규칙 (한글/영문)
    mapping_rules = {
        'control_code': ['통제코드', '코드', 'control code', 'code', 'control_code'],
        'control_name': ['통제명', '통제이름', '통제활동', 'control name', 'control', 'control_name'],
        'control_description': ['통제설명', '설명', '통제내용', 'description', 'control description'],
        'key_control': ['핵심통제', '핵심통제여부', 'key control', 'key'],
        'control_frequency': ['빈도', '통제빈도', 'frequency'],
        'control_type': ['통제유형', '유형', 'type', 'control type'],
        'control_nature': ['통제성격', '성격', 'nature'],
        'process_area': ['프로세스', '업무영역', 'process', 'process area'],
        'risk_description': ['위험', '위험설명', 'risk', 'risk description'],
        'risk_impact': ['영향', '위험영향', 'impact'],
        'risk_likelihood': ['발생가능성', '가능성', 'likelihood'],
        'population': ['모집단', 'population'],
        'population_completeness_check': ['완전성점검', '완전성확인', 'completeness'],
        'population_count': ['모집단수', '건수', 'count'],
        'test_procedure': ['테스트절차', '절차', 'procedure', 'test'],
        'control_owner': ['통제담당자', '담당자', 'owner'],
        'control_performer': ['수행자', 'performer'],
        'evidence_type': ['증적유형', '증적', 'evidence']
    }

    for db_field, keywords in mapping_rules.items():
        for idx, header in enumerate(headers):
            header_lower = str(header).lower().strip()
            if any(keyword.lower() in header_lower for keyword in keywords):
                mapping[db_field] = idx
                break

    return mapping

# RCM 삭제
@bp_rcm.route('/<int:rcm_id>/delete', methods=['POST'])
@admin_required
def rcm_delete(rcm_id):
    """RCM 삭제 (비활성화)"""
    user_info = get_user_info()

    try:
        db = get_db()
        db.execute('UPDATE ca_rcm SET is_active = ? WHERE rcm_id = ?', ('N', rcm_id))
        db.commit()

        log_user_activity(user_info, 'RCM_DELETE',
                        f'RCM 삭제 - RCM ID: {rcm_id}',
                        f'/rcm/{rcm_id}/delete', request.remote_addr,
                        request.headers.get('User-Agent'),
                        {'rcm_id': rcm_id})

        return jsonify({
            'success': True,
            'message': 'RCM이 삭제되었습니다.'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'삭제 중 오류가 발생했습니다: {str(e)}'}), 500

# RCM API - 상태 조회
@bp_rcm.route('/api/<int:rcm_id>/status')
@login_required
def rcm_status_api(rcm_id):
    """RCM 상태 조회 API"""
    user_info = get_user_info()

    if not has_rcm_access(user_info['user_id'], rcm_id):
        return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403

    rcm_info = get_rcm_info(rcm_id)
    if not rcm_info:
        return jsonify({'success': False, 'message': 'RCM을 찾을 수 없습니다.'}), 404

    rcm_details = get_rcm_details(rcm_id)

    return jsonify({
        'success': True,
        'rcm_id': rcm_id,
        'rcm_name': rcm_info['rcm_name'],
        'control_category': rcm_info['control_category'],
        'total_controls': len(rcm_details),
        'upload_date': rcm_info['upload_date'],
        'completion_date': rcm_info['completion_date']
    })
